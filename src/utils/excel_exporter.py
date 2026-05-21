"""Excel导出功能"""

import os
from datetime import datetime
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from src.config import Paths


def export_to_excel(records: List[Dict], filepath: str = None) -> str:
    """将仿真数据导出为Excel文件

    Args:
        records: 数据记录列表，每条记录含 time, sv, pv, u, disturbance, error
        filepath: 导出文件路径，为None时自动生成

    Returns:
        导出的文件路径
    """
    if filepath is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = os.path.join(Paths.HISTORY_DATA, f"export_{timestamp}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "仿真数据"

    # 标题行
    headers = ["时间 (s)", "SV 设定值", "PV 过程值", "控制量 u", "干扰", "误差"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    # 数据行
    data_font = Font(name="微软雅黑", size=10)
    data_align = Alignment(horizontal="center")
    for row_idx, record in enumerate(records, 2):
        values = [
            round(record.get("time", 0), 3),
            round(record.get("sv", 0), 3),
            round(record.get("pv", 0), 3),
            round(record.get("u", 0), 3),
            round(record.get("disturbance", 0), 3),
            round(record.get("error", 0), 3),
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

    # 列宽
    col_widths = [12, 14, 14, 14, 10, 12]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    wb.save(filepath)
    return filepath
